#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tạo file Excel "Nhập liệu báo cáo tháng" để cả Phòng cùng điền trên Google Drive.

    python tools/tao_file_nhap_lieu.py            # tạo NhapLieu-BaoCaoThang.xlsx

Cách dùng: chạy script -> upload file lên Google Drive -> mở bằng Google Sheets
(Drive tự chuyển đổi) -> chia sẻ quyền sửa cho các thành viên -> dán ID sheet vào
biến NHAP_LIEU_SHEET_ID của workflow. Từ đó ai phụ trách dự án nào tự điền dòng của mình.

Cần openpyxl (chỉ chạy trên máy khi cần tạo lại file mẫu, GitHub Action KHÔNG cần).
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "1B2A5B"
DO = "E03A3C"
NHAT = "EEF2F7"
VIEN = Border(*[Side(style="thin", color="C9D2E3")] * 4)

TRANG_THAI = ["Đang tăng trưởng", "Sắp go-live", "Cần chú ý", "Chờ đối tác"]
MUC_DO = ["GẤP", "Theo dõi"]
TT_MOC = ["", "Hoàn thành", "Đã đến hạn", "Trượt hạn"]

# Mã dự án PHẢI khớp key trong index.html — đừng đổi cột này.
DU_AN = [
    ("tripcare", "TripCARE", "Hằng"),
    ("spa", "SPA GoSafe", "Hằng"),
    ("bamboocare", "BambooCARE", "Hằng"),
    ("wbooking", "wBooking (HIO)", "Hằng"),
    ("flightdelay", "Flight Delay", "Hằng"),
    ("asahi", "Asahi (NPUV + APUV)", "Thư"),
    ("mobifone", "MobiFone (Cổng BH)", "Thư"),
    ("zalopay", "Zalopay (BHRVMH)", "Thư"),
    ("nnx", "NNX (Trợ cấp nằm viện)", "Thư"),
    ("plus-ew", "Plus & EW Honda", "Tùng"),
    ("viettelstore", "Viettel Store", "Tùng"),
    ("sungroup", "Sun Group (BH nhúng)", "Tùng"),
    ("khac", "Các dự án khác", "Tùng"),
]

TU_DONG = {"tripcare", "wbooking"}   # số liệu lấy thẳng từ Google Sheet doanh thu


def de_muc(ws, tieu_de, rong):
    for i, (ten, w) in enumerate(zip(tieu_de, rong), start=1):
        o = ws.cell(1, i, ten)
        o.font = Font(bold=True, color="FFFFFF", size=11)
        o.fill = PatternFill("solid", fgColor=NAVY)
        o.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        o.border = VIEN
        ws.column_dimensions[o.column_letter].width = w
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"


def to_o(ws, hang_dau=2):
    for hang in ws.iter_rows(min_row=hang_dau):
        for o in hang:
            o.alignment = Alignment(vertical="top", wrap_text=True)
            o.border = VIEN


def tab_huong_dan(wb):
    ws = wb.create_sheet("HƯỚNG DẪN", 0)
    ws.column_dimensions["A"].width = 118
    dong = [
        ("FILE NHẬP LIỆU BÁO CÁO THÁNG — PHÒNG THƯƠNG MẠI ĐIỆN TỬ", True),
        ("", False),
        ("Mỗi người phụ trách dự án nào thì tự điền dòng của dự án đó. Dashboard sẽ tự lấy số từ file này.", False),
        ("", False),
        ("1. Tab KỲ BÁO CÁO — sửa mã kỳ, nhãn và ngày chốt số mỗi khi bắt đầu tháng mới.", False),
        ("2. Tab DỰ ÁN — mỗi dự án một dòng: trạng thái, doanh thu, kết quả, kế hoạch.", False),
        ("3. Tab KHÓ KHĂN — việc cần lãnh đạo hỗ trợ (hiện ngay trang chính của dashboard).", False),
        ("4. Tab MỐC THỜI GIAN — các mốc go-live, hạn thanh toán, sự kiện trong kỳ.", False),
        ("", False),
        ("QUY TẮC QUAN TRỌNG:", True),
        ("• Cột 'Mã dự án' là khoá nối với dashboard — KHÔNG sửa, không xoá dòng, không đổi thứ tự.", False),
        ("• Ô nhiều ý: mỗi ý một dòng trong cùng ô (bấm Alt+Enter để xuống dòng). Không đánh số thứ tự.", False),
        ("• Doanh thu điền theo ĐƠN VỊ TỶ ĐỒNG, dùng dấu chấm thập phân (ví dụ 2.05 nghĩa là 2,05 tỷ).", False),
        ("• Dòng TripCARE và wBooking: BỎ TRỐNG các cột số — dashboard tự lấy từ Google Sheet doanh thu.", False),
        ("• Chưa có số liệu thì để trống, đừng điền số 0 (số 0 sẽ hiển thị là 'đã phát sinh 0 đồng').", False),
        ("", False),
        ("Điền xong không cần làm gì thêm: dashboard tự cập nhật 2 lần/ngày (7h và 17h).", False),
    ]
    for i, (txt, dam) in enumerate(dong, start=1):
        o = ws.cell(i, 1, txt)
        o.font = Font(bold=dam, size=13 if i == 1 else 11, color=NAVY if dam else "16203A")
        o.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[i].height = 22


def tab_ky(wb):
    ws = wb.create_sheet("KỲ BÁO CÁO")
    de_muc(ws, ["Trường", "Giá trị", "Giải thích"], [26, 34, 62])
    dong = [
        ("Mã kỳ", "2026-09", "Định dạng NĂM-THÁNG, ví dụ 2026-10 cho tháng 10"),
        ("Nhãn hiển thị", "Tháng 09/2026", "Tên hiện trên ô chọn kỳ của dashboard"),
        ("Khoảng thời gian", "01–30/09/2026", ""),
        ("Chốt số đến ngày", "30/09/2026", "Ngày số liệu được chốt"),
        ("Mục tiêu năm (tỷ)", 100, "Mục tiêu doanh thu cả năm của Phòng"),
        ("Lưu ý chung", "", "Câu lưu ý hiện kèm bảng doanh thu (để trống nếu không có)"),
    ]
    for i, (a, b, c) in enumerate(dong, start=2):
        ws.cell(i, 1, a).font = Font(bold=True)
        ws.cell(i, 2, b)
        ws.cell(i, 3, c).font = Font(italic=True, color="5B6577")
    to_o(ws)


def tab_du_an(wb):
    ws = wb.create_sheet("DỰ ÁN")
    de_muc(ws, ["Mã dự án\n(không sửa)", "Tên dự án", "Phụ trách", "Trạng thái",
                "DT tháng\n(tỷ)", "Lũy kế\n(tỷ)", "% KPI năm", "Cùng kỳ 2025\n(tỷ)",
                "Ghi chú (1–2 câu tóm tắt cho lãnh đạo)",
                "Kết quả trong kỳ\n(mỗi ý một dòng)", "Kế hoạch tiếp theo\n(mỗi ý một dòng)"],
            [16, 22, 11, 16, 10, 10, 11, 13, 46, 52, 52])
    for i, (ma, ten, nguoi) in enumerate(DU_AN, start=2):
        ws.cell(i, 1, ma).font = Font(bold=True, color=NAVY)
        ws.cell(i, 2, ten)
        ws.cell(i, 3, nguoi)
        if ma in TU_DONG:
            o = ws.cell(i, 5, "(tự động)")
            o.font = Font(italic=True, color="5B6577")
            o.fill = PatternFill("solid", fgColor=NHAT)
            for cot in (6, 7, 8):
                ws.cell(i, cot).fill = PatternFill("solid", fgColor=NHAT)
        ws.row_dimensions[i].height = 58
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(TRANG_THAI), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("D2:D%d" % (len(DU_AN) + 1))
    to_o(ws)


def tab_kho_khan(wb):
    ws = wb.create_sheet("KHÓ KHĂN")
    de_muc(ws, ["Mức độ", "Dự án", "Khó khăn / vướng mắc", "Đề xuất lãnh đạo hỗ trợ"],
           [12, 22, 62, 62])
    for i in range(2, 14):
        ws.row_dimensions[i].height = 46
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(MUC_DO), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("A2:A30")
    to_o(ws)


def tab_moc(wb):
    ws = wb.create_sheet("MỐC THỜI GIAN")
    de_muc(ws, ["Ngày / thời điểm", "Dự án", "Nội dung mốc", "Quan trọng?\n(x)", "Trạng thái"],
           [18, 22, 62, 12, 18])
    for i in range(2, 26):
        ws.row_dimensions[i].height = 30
    dv1 = DataValidation(type="list", formula1='"%s"' % ",".join(TT_MOC[1:]), allow_blank=True)
    ws.add_data_validation(dv1)
    dv1.add("E2:E40")
    dv2 = DataValidation(type="list", formula1='"x"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add("D2:D40")
    to_o(ws)


def tab_vi_du(wb):
    """Một tab chỉ để xem cho biết cách điền — script KHÔNG đọc tab này."""
    ws = wb.create_sheet("VÍ DỤ (chỉ để xem)")
    de_muc(ws, ["Mã dự án", "Tên dự án", "Phụ trách", "Trạng thái", "DT tháng\n(tỷ)",
                "Lũy kế\n(tỷ)", "% KPI năm", "Cùng kỳ 2025\n(tỷ)", "Ghi chú",
                "Kết quả trong kỳ", "Kế hoạch tiếp theo"],
            [14, 20, 11, 15, 10, 10, 11, 13, 44, 50, 50])
    vi_du = [
        ("asahi", "Asahi (NPUV + APUV)", "Thư", "Cần chú ý", 2.05, 13.03, 46.12, None,
         "T7 đạt 2,05 tỷ; lũy kế 7T 13,03 tỷ = 46,12% kế hoạch năm (28 tỷ)",
         "T7: 2,05 tỷ · lũy kế 7T 13,03 tỷ\n"
         "Nâng cấp landing page NPUV, xây landing page thu lead cho ALCV\n"
         "Hoàn thành bản chào phí BH tai nạn cho giải chạy Asahi tài trợ",
         "Thống nhất phương án điều chỉnh APUV với Ban CN & Actuary\n"
         "Hoàn thiện chương trình khuyến khích bán hàng cho TSR của ALCV"),
        ("viettelstore", "Viettel Store", "Tùng", "Cần chú ý", None, None, None, None,
         "VST chưa chốt ngày go-live ADLD; BH '1 đổi 1' đặt mục tiêu go-live T8 kịp iPhone 18",
         "Test xong luồng API mới chương trình ADLD với IT hai bên\n"
         "Hoàn thiện bộ hợp đồng, phụ lục, API cho chương trình '1 đổi 1'",
         "Thúc đẩy VST chốt ngày go-live ADLD\n"
         "Kết nối API & go-live BH '1 đổi 1' trong tháng"),
    ]
    for i, hang in enumerate(vi_du, start=2):
        for j, gt in enumerate(hang, start=1):
            ws.cell(i, j, gt)
        ws.row_dimensions[i].height = 74
    ws.cell(5, 1, "Chú ý: mỗi ý một dòng trong cùng ô (Alt+Enter). Ô nào chưa có số thì để TRỐNG.").font = \
        Font(italic=True, color=DO, bold=True)
    to_o(ws, hang_dau=2)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    tab_huong_dan(wb)
    tab_ky(wb)
    tab_du_an(wb)
    tab_kho_khan(wb)
    tab_moc(wb)
    tab_vi_du(wb)
    duong_dan = Path(__file__).resolve().parent.parent / "NhapLieu-BaoCaoThang.xlsx"
    wb.save(duong_dan)
    print("Đã tạo %s" % duong_dan)


if __name__ == "__main__":
    main()
